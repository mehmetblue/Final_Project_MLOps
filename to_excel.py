import pandas as pd
from openpyxl.utils import get_column_letter
from io import BytesIO

def to_excel_auto_width(df):
    output = BytesIO()
    # Convert DataFrame to an Excel file
    with pd.ExcelWriter(output, engine='openpyxl', mode='w') as writer:
        df.to_excel(writer, index=False, sheet_name='Sheet1')
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']
        
        # Adjust column widths
        for col_num, column in enumerate(worksheet.iter_cols(min_row=1, max_row=worksheet.max_row), start=1):
            max_length = 0
            for cell in column:
                cell_length = len(str(cell.value))
                max_length = max(max_length, cell_length)
            
            # Set the column width in Excel, adding a bit extra space by using +2
            worksheet.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    output.seek(0)  # Reset the file pointer to the beginning
    return output
